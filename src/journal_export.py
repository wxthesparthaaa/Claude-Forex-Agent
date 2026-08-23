"""
Builds an Excel workbook from the trade journal for weekend review --
generated on demand from trade_journal.json (the source of truth,
already GitHub-synced), not maintained as a separate persisted binary
file. Always reflects the current journal, no separate sync path to
keep consistent or risk going stale.
"""
from __future__ import annotations

from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

COLUMNS = [
    "Trade ID", "Instrument", "Direction", "Units", "Entry Price", "Stop Loss", "Take Profit",
    "Confidence %", "Breadth", "RSI", "Candlestick", "News", "Status", "R Multiple",
    "Opened At", "Closed At", "Hours Held", "Exit Price", "Realized P&L", "Currency",
    "Experiment", "Parent Trade ID", "Rationale",
]


def _parse_iso(ts: str) -> datetime:
    """Handles both our own opened_at format (isoformat(), 6-digit
    microseconds) and OANDA's closed_at format (nanosecond fraction --
    or sometimes none at all -- plus a trailing "Z"), which
    datetime.fromisoformat() can't parse directly without the "Z" swap.
    Python's fromisoformat() (3.11+) accepts a fractional-seconds part
    of any length itself, so no manual truncation is needed -- an
    earlier version here sliced to a fixed 26 characters assuming a
    9-digit fraction was always present, which broke on a real OANDA
    timestamp with no fractional seconds at all."""
    if ts.endswith("Z"):
        ts = ts[:-1] + "+00:00"
    return datetime.fromisoformat(ts)


def _r_multiple(entry: dict) -> float | None:
    """P&L expressed in units of the original risk (R) -- lets you
    compare trade quality independent of position size, the same metric
    the backtest engine uses."""
    if entry.get("exit_price") is None:
        return None
    risk = abs(float(entry["entry_price"]) - float(entry["stop_loss"]))
    if risk == 0:
        return None
    direction_sign = 1 if entry["direction"] == "LONG" else -1
    return round(direction_sign * (float(entry["exit_price"]) - float(entry["entry_price"])) / risk, 2)


def _hours_held(entry: dict) -> float | None:
    if not entry.get("closed_at"):
        return None
    try:
        opened = _parse_iso(entry["opened_at"])
        closed = _parse_iso(entry["closed_at"])
        return round((closed - opened).total_seconds() / 3600, 2)
    except (KeyError, ValueError):
        return None


def build_journal_workbook(entries: list) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Trade Journal"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for entry in entries:
        components = entry.get("confidence_components") or {}
        ws.append([
            entry.get("trade_id"), entry.get("instrument"), entry.get("direction"), entry.get("units"),
            entry.get("entry_price"), entry.get("stop_loss"), entry.get("take_profit"),
            entry.get("confidence_pct"),
            components.get("breadth"), components.get("rsi"), components.get("candlestick"), components.get("news"),
            entry.get("status"), _r_multiple(entry),
            entry.get("opened_at"), entry.get("closed_at"), _hours_held(entry),
            entry.get("exit_price"), entry.get("realized_pnl"), entry.get("account_currency"),
            entry.get("experiment_tag") or "", entry.get("parent_trade_id") or "",
            " | ".join(entry.get("rationale", [])),
        ])

    for i, _ in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.column_dimensions[get_column_letter(len(COLUMNS))].width = 60  # Rationale

    return wb
